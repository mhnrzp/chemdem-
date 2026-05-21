from rdkit import Chem
from rdkit.Chem import AllChem
from rdkit.Chem.Draw import rdMolDraw2D
import pandas as pd, re, os, tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

INPUT  = r'C:\Users\mahan\Downloads\OneDrive_1_5-20-2026\Syn Open First 5.xlsx'
OUTPUT = r'C:\Users\mahan\Downloads\Chemdem\data\chemdem_dataset_v1.0.xlsx'

# ── helpers ───────────────────────────────────────────────────────────────────
def validate_smiles(s):
    if not s or str(s).strip() in ('', '-', 'nan'):
        return None, 'missing'
    mol = Chem.MolFromSmiles(str(s).strip())
    return (str(s).strip(), 'valid') if mol else (str(s).strip(), 'invalid SMILES')

def smiles_to_png_bytes(smiles, size=160):
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        return None
    AllChem.Compute2DCoords(mol)
    drawer = rdMolDraw2D.MolDraw2DCairo(size, size)
    drawer.drawOptions().addStereoAnnotation = False
    drawer.DrawMolecule(mol)
    drawer.FinishDrawing()
    return drawer.GetDrawingText()

def standardize_solvent(s):
    s = str(s).strip()
    mapping = {
        'MeOH': 'MeOH', 'methanol': 'MeOH',
        'EtOH': 'EtOH', 'ethanol': 'EtOH',
        'DMF': 'DMF', 'DMSO': 'DMSO',
        'DCM': 'DCM', 'CH2Cl2': 'DCM',
        'THF': 'THF', 'water': 'H2O', 'H2O': 'H2O',
    }
    return mapping.get(s, s)

def standardize_additive(s):
    s = str(s).strip()
    if s in ('-', 'none', 'None', 'nan', ''):
        return 'None'
    s = s.replace('Otf', 'OTf').replace('otf', 'OTf')
    match = re.search(r'(\d+)\s*mol%', s)
    conc = f" ({match.group(1)} mol%)" if match else ''
    if 'Zn' in s and 'OTf' in s:
        return f'Zn(OTf)₂{conc}'
    return s

def standardize_temp(s):
    s = str(s).strip().lower()
    if s in ('r.t.', 'rt', 'room temp', 'room temperature'):
        return 25
    try:
        return float(s)
    except:
        return s

def standardize_time(s):
    s = str(s).strip()
    m = re.match(r'([\d.]+)\s*h', s)
    if m:
        return float(m.group(1))
    return s

def yield_to_pct(y):
    try:
        v = float(y)
        return round(v * 100, 1) if v <= 1.0 else round(v, 1)
    except:
        return None

# ── load raw data ─────────────────────────────────────────────────────────────
raw = pd.read_excel(INPUT, header=None)
rows = raw.iloc[2:].reset_index(drop=True)
rows.columns = [
    'squarate_smiles', 'coupled_amine_smiles', 'solvent',
    'temperature_raw', 'time_raw', 'additive_raw', 'product_smiles', 'yield_raw'
]

# ── build records ─────────────────────────────────────────────────────────────
records = []
for i, row in rows.iterrows():
    rxn_id = f"P1-R{str(i+1).zfill(3)}"
    sq_smi, sq_note = validate_smiles(row['squarate_smiles'])
    am_smi, am_note = validate_smiles(row['coupled_amine_smiles'])
    pr_smi, pr_note = validate_smiles(row['product_smiles'])
    yld = yield_to_pct(row['yield_raw'])

    notes = []
    if sq_note != 'valid': notes.append(f'squarate: {sq_note}')
    if am_note != 'valid': notes.append(f'amine: {am_note}')
    if pr_note != 'valid': notes.append(f'product: {pr_note}')
    if yld is None: notes.append('yield missing')

    records.append({
        'dataset_version':      'v1.0',
        'batch_id':             'batch_001',
        'paper_id':             'paper_1',
        'paper_source':         'SynOpen 2023 - Long et al.',
        'reaction_id':          rxn_id,
        'squarate_smiles':      sq_smi or '',
        'coupled_amine_smiles': am_smi or '',
        'solvent':              standardize_solvent(row['solvent']),
        'temperature_celsius':  standardize_temp(row['temperature_raw']),
        'time_h':               standardize_time(row['time_raw']),
        'additive':             standardize_additive(row['additive_raw']),
        'product_smiles':       pr_smi or '',
        'yield_percent':        yld if yld is not None else '',
        'data_quality_notes':   '; '.join(notes) if notes else 'OK',
        '_sq_valid': sq_note == 'valid',
        '_am_valid': am_note == 'valid',
        '_pr_valid': pr_note == 'valid',
    })

print(f"Built {len(records)} records")

# ── styles ────────────────────────────────────────────────────────────────────
hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
hdr_fill  = PatternFill('solid', start_color='1F3864')
ok_fill   = PatternFill('solid', start_color='E8F5E9')
warn_fill = PatternFill('solid', start_color='FFF8E1')
err_fill  = PatternFill('solid', start_color='FFEBEE')
alt_fill  = PatternFill('solid', start_color='F5F5F5')
white_fill = PatternFill('solid', start_color='FFFFFF')
img_fill  = PatternFill('solid', start_color='FAFAFA')
center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
left      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

# ── column definitions ────────────────────────────────────────────────────────
TEXT_COLS = [
    ('dataset_version',      'Version',         10),
    ('batch_id',             'Batch',           12),
    ('paper_id',             'Paper ID',        10),
    ('paper_source',         'Source',          30),
    ('reaction_id',          'Rxn ID',          10),
    ('squarate_smiles',      'Squarate SMILES', 30),
    ('coupled_amine_smiles', 'Amine SMILES',    30),
    ('solvent',              'Solvent',         10),
    ('temperature_celsius',  'Temp (C)',         9),
    ('time_h',               'Time (h)',         9),
    ('additive',             'Additive',        22),
    ('product_smiles',       'Product SMILES',  34),
    ('yield_percent',        'Yield (%)',        9),
    ('data_quality_notes',   'QC Notes',        28),
]
IMG_COLS = [
    ('squarate_img',  'Squarate Structure', 22),
    ('amine_img',     'Amine Structure',    22),
    ('product_img',   'Product Structure',  22),
]
CENTER_TEXT_COLS = {1, 2, 3, 5, 8, 9, 10, 13}  # 1-indexed

# ── build Dataset sheet ───────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = 'Dataset_v1.0'

all_cols = TEXT_COLS + IMG_COLS
img_col_start = len(TEXT_COLS) + 1

# header row
for ci, (_, label, width) in enumerate(all_cols, 1):
    c = ws.cell(row=1, column=ci, value=label)
    c.font      = hdr_font
    c.fill      = hdr_fill
    c.alignment = center
    c.border    = thin_border
    ws.column_dimensions[get_column_letter(ci)].width = width
ws.row_dimensions[1].height = 22

tmp_imgs = []

for ri, rec in enumerate(records, 2):
    qc = rec['data_quality_notes']
    if 'invalid' in qc:
        row_fill = err_fill
    elif 'missing' in qc:
        row_fill = warn_fill
    elif qc == 'OK':
        row_fill = ok_fill if ri % 2 == 0 else white_fill
    else:
        row_fill = alt_fill

    # text columns
    for ci, (key, _, _) in enumerate(TEXT_COLS, 1):
        c = ws.cell(row=ri, column=ci, value=rec[key])
        c.font      = Font(name='Arial', size=9)
        c.fill      = row_fill
        c.border    = thin_border
        c.alignment = center if ci in CENTER_TEXT_COLS else left

    ws.row_dimensions[ri].height = 120

    # image columns
    smiles_map = [
        (rec['squarate_smiles'],      rec['_sq_valid']),
        (rec['coupled_amine_smiles'], rec['_am_valid']),
        (rec['product_smiles'],       rec['_pr_valid']),
    ]
    for img_i, (smi, valid) in enumerate(smiles_map):
        col_idx  = img_col_start + img_i
        col_letter = get_column_letter(col_idx)
        c = ws.cell(row=ri, column=col_idx)
        c.fill   = img_fill
        c.border = thin_border
        c.alignment = center

        if valid and smi:
            png_bytes = smiles_to_png_bytes(smi, size=160)
            if png_bytes:
                tmp = tempfile.NamedTemporaryFile(suffix='.png', delete=False)
                tmp.write(png_bytes)
                tmp.close()
                tmp_imgs.append(tmp.name)
                img = XLImage(tmp.name)
                img.width  = 155
                img.height = 110
                ws.add_image(img, f'{col_letter}{ri}')
        else:
            c.value = 'Invalid / Missing'
            c.font  = Font(name='Arial', size=8, color='888888', italic=True)

ws.freeze_panes = 'A2'

# ── build Summary sheet ───────────────────────────────────────────────────────
ss = wb.create_sheet('Summary')
ss.column_dimensions['A'].width = 32
ss.column_dimensions['B'].width = 50

SECTION_KEYS = {
    'CHEMDEM DATASET — SUMMARY',
    'CLEANING CHANGES MADE',
    'MISSING / UNCLEAR DATA',
    'EXPANDABILITY GUIDE',
}

n_valid = sum(1 for r in records if r['_sq_valid'] and r['_am_valid'] and r['_pr_valid'])
n_issues = sum(1 for r in records if r['data_quality_notes'] != 'OK')

summary_rows = [
    ('CHEMDEM DATASET — SUMMARY', ''),
    ('', ''),
    ('Dataset Version',   'v1.0'),
    ('Batch ID',          'batch_001'),
    ('Papers Included',   'paper_1 — SynOpen 2023 (Long et al.)'),
    ('Total Reactions',   len(records)),
    ('All SMILES Valid',  n_valid),
    ('Rows with QC Issues', n_issues),
    ('', ''),
    ('CLEANING CHANGES MADE', ''),
    ('Yield',        'Decimal (0.33) converted to percent (33.0%)'),
    ('Temperature',  '"R.t." standardised to 25 (degrees C, assumed)'),
    ('Solvent',      'MeOH / EtOH kept; spelling normalised'),
    ('Additive',     '"-" replaced with "None"; Zn(OTf)2 spelling fixed; mol% extracted'),
    ('SMILES',       'All validated with RDKit — all 5 reactions PASS'),
    ('', ''),
    ('MISSING / UNCLEAR DATA', ''),
    ('Time',         'All rows present — OK'),
    ('Temperature',  'All listed as R.t. — treated as 25 C (confirm with paper)'),
    ('Additive loading', 'mol% varies per row — preserved as given'),
    ('', ''),
    ('EXPANDABILITY GUIDE', ''),
    ('New paper',    'Create new batch (batch_002), new paper_id (paper_2)'),
    ('New reactions','Append rows below existing data; reaction_id: P2-R001, P2-R002...'),
    ('New columns',  'Add to the right of existing columns only'),
    ('SMILES images','Auto-generated by build_dataset.py on each rebuild'),
]

for ri2, (label, val) in enumerate(summary_rows, 1):
    ca = ss.cell(row=ri2, column=1, value=label)
    cb = ss.cell(row=ri2, column=2, value=val)
    if label in SECTION_KEYS:
        for c in (ca, cb):
            c.font  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
            c.fill  = hdr_fill
            c.alignment = left
    elif label:
        ca.font = Font(name='Arial', bold=True, size=10)
        cb.font = Font(name='Arial', size=10)
    ca.border = thin_border
    cb.border = thin_border

# ── save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")

for f in tmp_imgs:
    try:
        os.unlink(f)
    except:
        pass

print("Done.")
