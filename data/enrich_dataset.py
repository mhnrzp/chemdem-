"""
enrich_dataset.py  ─  Phase 4: PubChem-powered dataset enrichment
────────────────────────────────────────────────────────────────────
Reads  : chemdem_dataset_v1.0.xlsx  (5 Excel rows)
         table1_reactions.csv       (51 encoded rows, Papers 1-3)
         ChemInventory-Export.xlsx  (276 lecturer compounds)
Calls  : PubChem REST API (free, no key)
Writes : chemdem_dataset_v1.1.xlsx  (full enriched dataset)
"""

import pandas as pd, requests, time, re, json, os, tempfile
from urllib.parse import quote
from rdkit import Chem
from rdkit.Chem import AllChem, rdMolDescriptors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE    = r'C:\Users\mahan\Downloads\Chemdem\data'
LIB_XL  = r'C:\Users\mahan\Downloads\ChemInventory-Export.xlsx'
IN_XL   = os.path.join(BASE, 'chemdem_dataset_v1.0.xlsx')
IN_CSV  = os.path.join(BASE, 'table1_reactions.csv')
OUT_XL  = os.path.join(BASE, 'chemdem_dataset_v1.1.xlsx')
CACHE_F = os.path.join(BASE, 'pubchem_cache.json')

# ── SMILES master lookup ───────────────────────────────────────────────────────
# Maps (amine_type, substituent, position) → compound info
AMINE_MAP = {
    # Benzylamine series
    ('benzylamine',  'none',  'none'):  {'name':'Benzylamine',               'smiles':'NCC1=CC=CC=C1',           'cas':'100-46-9'},
    ('benzylamine',  'Cl',    'para'):  {'name':'4-Chlorobenzylamine',        'smiles':'NCC1=CC=C(Cl)C=C1',       'cas':'104-86-9'},
    ('benzylamine',  'Cl',    'meta'):  {'name':'3-Chlorobenzylamine',        'smiles':'NCC1=CC=CC(Cl)=C1',       'cas':'4152-90-3'},
    ('benzylamine',  'Br',    'para'):  {'name':'4-Bromobenzylamine',         'smiles':'NCC1=CC=C(Br)C=C1',       'cas':'3959-07-7'},
    ('benzylamine',  'Br',    'meta'):  {'name':'3-Bromobenzylamine',         'smiles':'NCC1=CC=CC(Br)=C1',       'cas':'16642-79-8'},
    # Aniline series
    ('aniline',      'none',  'none'):  {'name':'Aniline',                    'smiles':'NC1=CC=CC=C1',            'cas':'62-53-3'},
    ('aniline',      'F',     'para'):  {'name':'4-Fluoroaniline',            'smiles':'NC1=CC=C(F)C=C1',         'cas':'371-40-4'},
    ('aniline',      'F',     'meta'):  {'name':'3-Fluoroaniline',            'smiles':'NC1=CC=CC(F)=C1',         'cas':'372-19-0'},
    ('aniline',      'F',     'ortho'): {'name':'2-Fluoroaniline',            'smiles':'NC1=CC=CC=C1F',           'cas':'348-54-9'},
    ('aniline',      'Cl',    'para'):  {'name':'4-Chloroaniline',            'smiles':'NC1=CC=C(Cl)C=C1',        'cas':'106-47-8'},
    ('aniline',      'Cl',    'meta'):  {'name':'3-Chloroaniline',            'smiles':'NC1=CC=CC(Cl)=C1',        'cas':'108-42-9'},
    ('aniline',      'Cl',    'ortho'): {'name':'2-Chloroaniline',            'smiles':'NC1=CC=CC=C1Cl',          'cas':'95-51-2'},
    ('aniline',      'Br',    'para'):  {'name':'4-Bromoaniline',             'smiles':'NC1=CC=C(Br)C=C1',        'cas':'106-40-1'},
    ('aniline',      'Br',    'meta'):  {'name':'3-Bromoaniline',             'smiles':'NC1=CC=CC(Br)=C1',        'cas':'591-19-5'},
    ('aniline',      'Br',    'ortho'): {'name':'2-Bromoaniline',             'smiles':'NC1=CC=CC=C1Br',          'cas':'615-36-1'},
    ('aniline',      'OH',    'para'):  {'name':'4-Aminophenol',              'smiles':'NC1=CC=C(O)C=C1',         'cas':'123-30-8'},
    ('aniline',      'OH',    'meta'):  {'name':'3-Aminophenol',              'smiles':'NC1=CC=CC(O)=C1',         'cas':'591-27-5'},
    ('aniline',      'OH',    'ortho'): {'name':'2-Aminophenol',              'smiles':'NC1=CC=CC=C1O',           'cas':'95-55-6'},
    ('aniline',      'OMe',   'para'):  {'name':'4-Methoxyaniline',           'smiles':'NC1=CC=C(OC)C=C1',        'cas':'104-94-9'},
    ('aniline',      'OMe',   'meta'):  {'name':'3-Methoxyaniline',           'smiles':'NC1=CC=CC(OC)=C1',        'cas':'536-90-3'},
    ('aniline',      'OMe',   'ortho'): {'name':'2-Methoxyaniline',           'smiles':'NC1=CC=CC=C1OC',          'cas':'90-04-0'},
    ('aniline',      'CF3',   'para'):  {'name':'4-(Trifluoromethyl)aniline', 'smiles':'NC1=CC=C(C(F)(F)F)C=C1', 'cas':'455-14-1'},
    ('aniline',      'CF3',   'meta'):  {'name':'3-(Trifluoromethyl)aniline', 'smiles':'NC1=CC=CC(C(F)(F)F)=C1', 'cas':'98-16-8'},
    ('aniline',      'Me',    'para'):  {'name':'4-Toluidine',                'smiles':'NC1=CC=C(C)C=C1',         'cas':'106-49-0'},
    ('aniline',      'Me',    'meta'):  {'name':'3-Toluidine',                'smiles':'NC1=CC=CC(C)=C1',         'cas':'108-44-1'},
    ('aniline',      'Me',    'ortho'): {'name':'2-Toluidine',                'smiles':'NC1=CC=CC=C1C',           'cas':'95-53-4'},
    # Heterocyclic amines
    ('heterocyclic', 'pyrrolidine',        'none'): {'name':'Pyrrolidine',            'smiles':'C1CCNC1',                 'cas':'123-75-1'},
    ('heterocyclic', 'azetidine',          'none'): {'name':'Azetidine',              'smiles':'C1CNC1',                  'cas':'503-29-7'},
    ('heterocyclic', 'piperidine',         'none'): {'name':'Piperidine',             'smiles':'C1CCNCC1',                'cas':'110-89-4'},
    ('heterocyclic', 'azepane',            'none'): {'name':'Azepane',                'smiles':'C1CCNCCC1',               'cas':'111-49-9'},
    ('heterocyclic', 'morpholine',         'none'): {'name':'Morpholine',             'smiles':'C1COCCN1',                'cas':'110-91-8'},
    ('heterocyclic', 'thiomorpholine',     'none'): {'name':'Thiomorpholine',         'smiles':'C1CSCNC1',                'cas':'123-90-0'},
    ('heterocyclic', 'thiomorpholine-1,1-dioxide','none'):{'name':'Thiomorpholine 1,1-dioxide','smiles':'O=S1(=O)CCNCC1','cas':'39093-93-1'},
    ('heterocyclic', 'N-Boc-piperazine',   'none'): {'name':'N-Boc-piperazine',       'smiles':'CC(C)(C)OC(=O)N1CCNCC1', 'cas':'57260-71-6'},
    ('heterocyclic', 'piperazine',         'none'): {'name':'Piperazine',             'smiles':'C1CNCCN1',                'cas':'110-85-0'},
    ('heterocyclic', 'thiazolidine',       'none'): {'name':'Thiazolidine',           'smiles':'C1CSCN1',                 'cas':'504-78-9'},
    # Ortho-halide benzylamines
    ('benzylamine',  'Cl',    'ortho'): {'name':'2-Chlorobenzylamine',       'smiles':'NCC1=CC=CC=C1Cl',         'cas':'89-92-9'},
    ('benzylamine',  'Br',    'ortho'): {'name':'2-Bromobenzylamine',        'smiles':'NCC1=CC=CC=C1Br',         'cas':'3959-08-8'},
    # More heterocyclics (Paper 2 extras)
    ('heterocyclic', 'dimethylmorpholine',      'none'): {'name':'2,6-Dimethylmorpholine',   'smiles':'CC1COCCNC1C',            'cas':'141-91-3'},
    ('heterocyclic', 'boc_dimethylpiperazine',  'none'): {'name':'N-Boc-2,5-dimethylpiperazine','smiles':'CC1CN(C(=O)OC(C)(C)C)CC(C)N1', 'cas':'169447-70-5'},
    ('heterocyclic', 'bromo_tetrahydroquinoline','none'): {'name':'6-Bromo-1,2,3,4-tetrahydroquinoline','smiles':'BrC1=CC=C2CCCNC2=C1',   'cas':'40013-86-3'},
    ('heterocyclic', 'thiomorpholine_dioxide',  'none'): {'name':'Thiomorpholine 1,1-dioxide', 'smiles':'O=S1(=O)CCNCC1',         'cas':'39093-93-1'},
    ('heterocyclic', 'boc_piperazine',          'none'): {'name':'N-Boc-piperazine',          'smiles':'CC(C)(C)OC(=O)N1CCNCC1', 'cas':'57260-71-6'},
    # Paper 3 bis-squaramide / fused-bicycle starters (amine component only)
    ('bis_squaramide_methylation', 'aniline',      'none'): {'name':'Aniline',       'smiles':'NC1=CC=CC=C1',    'cas':'62-53-3'},
    ('bis_squaramide_methylation', 'benzylamine',  'none'): {'name':'Benzylamine',   'smiles':'NCC1=CC=CC=C1',   'cas':'100-46-9'},
    ('bis_squaramide_methylation', 'bromoaniline', 'none'): {'name':'4-Bromoaniline','smiles':'NC1=CC=C(Br)C=C1','cas':'106-40-1'},
    ('bis_squaramide_benzylation', 'bromoaniline', 'para'): {'name':'4-Bromoaniline','smiles':'NC1=CC=C(Br)C=C1','cas':'106-40-1'},
    ('fused_bicycle', 'diisopropyl_diamine', 'none'): {'name':'N,N-Diisopropylethylenediamine','smiles':'CC(N)CNC(C)C',          'cas':'121-05-1'},
    ('fused_bicycle', 'dimethyl_diamine',    'none'): {'name':'N,N-Dimethylethylenediamine',  'smiles':'CNCCNC',                 'cas':'110-70-3'},
}

SQUARATE_SMILES = 'O=C1C(OCC)=C(OCC)C1=O'
SQUARATE_CAS    = '5231-87-8'
SQUARATE_NAME   = 'Diethyl squarate'

# RDKit SMARTS for squarate monoamide coupling
REACTION_SMARTS = AllChem.ReactionFromSmarts(
    '[NH2:1].[C:2](=O)[C:3](OCC)=C([C:4]=O)[C:5]>>[NH:1][C:3](=O)[C:2]=[C:4][C:5]=O'
)

# ── RDKit helpers ─────────────────────────────────────────────────────────────
def canonicalise(smi):
    if not smi or str(smi).strip() in ('', 'nan', '-'): return None
    mol = Chem.MolFromSmiles(str(smi).strip())
    return Chem.MolToSmiles(mol) if mol else None

def formula_from_smiles(smi):
    mol = Chem.MolFromSmiles(str(smi).strip()) if smi else None
    return rdMolDescriptors.CalcMolFormula(mol) if mol else None

def smiles_match(s1, s2):
    c1, c2 = canonicalise(s1), canonicalise(s2)
    return c1 and c2 and c1 == c2

def compute_product(amine_smi, squarate_smi=SQUARATE_SMILES):
    """Use RDKit SMARTS to predict product SMILES."""
    am  = Chem.MolFromSmiles(amine_smi)
    sq  = Chem.MolFromSmiles(squarate_smi)
    if am is None or sq is None: return None
    try:
        products = REACTION_SMARTS.RunReactants((am, sq))
        if products:
            p_smi = Chem.MolToSmiles(products[0][0])
            # Validate the product
            if Chem.MolFromSmiles(p_smi): return p_smi
    except Exception: pass
    return None

def cas_pattern(synonym):
    return bool(re.match(r'^\d{2,7}-\d{2}-\d$', synonym.strip()))

# ── PubChem API ───────────────────────────────────────────────────────────────
PUBCHEM_BASE = 'https://pubchem.ncbi.nlm.nih.gov/rest/pug'
HEADERS      = {'Accept': 'application/json'}

def pubchem_by_smiles(smiles, cache):
    canon = canonicalise(smiles)
    if not canon: return {}
    if canon in cache: return cache[canon]

    result = {}
    try:
        enc = quote(canon, safe='')
        url = f'{PUBCHEM_BASE}/compound/smiles/{enc}/property/IUPACName,MolecularFormula,CanonicalSMILES,MolecularWeight/JSON'
        r = requests.get(url, headers=HEADERS, timeout=12)
        if r.status_code == 200:
            props = r.json()['PropertyTable']['Properties'][0]
            cid   = props.get('CID')
            result = {
                'cid':            cid,
                'pubchem_iupac':  props.get('IUPACName', ''),
                'pubchem_formula':props.get('MolecularFormula', ''),
                'pubchem_smiles': props.get('ConnectivitySMILES', props.get('CanonicalSMILES', props.get('IsomericSMILES', ''))),
                'pubchem_mw':     props.get('MolecularWeight', ''),
                'pubchem_cas':    '',
            }
            # Fetch CAS from synonyms
            if cid:
                time.sleep(0.25)
                syn_url = f'{PUBCHEM_BASE}/compound/cid/{cid}/synonyms/JSON'
                sr = requests.get(syn_url, headers=HEADERS, timeout=10)
                if sr.status_code == 200:
                    syns = sr.json()['InformationList']['Information'][0].get('Synonym', [])
                    cas_hits = [s for s in syns if cas_pattern(s)]
                    result['pubchem_cas'] = cas_hits[0] if cas_hits else ''
        elif r.status_code == 404:
            result = {'cid': None, 'pubchem_iupac': 'NOT FOUND', 'pubchem_formula': '',
                      'pubchem_smiles': '', 'pubchem_mw': '', 'pubchem_cas': ''}
        else:
            result = {'cid': None, 'pubchem_iupac': f'HTTP {r.status_code}', 'pubchem_formula': '',
                      'pubchem_smiles': '', 'pubchem_mw': '', 'pubchem_cas': ''}
    except Exception as ex:
        result = {'cid': None, 'pubchem_iupac': f'ERROR: {ex}', 'pubchem_formula': '',
                  'pubchem_smiles': '', 'pubchem_mw': '', 'pubchem_cas': ''}
    cache[canon] = result
    time.sleep(0.35)   # respect PubChem rate limit
    return result

# ── Load ChemInventory ────────────────────────────────────────────────────────
def load_library():
    df = pd.read_excel(LIB_XL)
    df.columns = [c.strip() for c in df.columns]
    lib = {}
    for _, row in df.iterrows():
        cas = str(row.get('Substance CAS', '')).strip()
        smi = str(row.get('SMILES', '')).strip()
        frm = str(row.get('Molecular Formula', '')).strip()
        nm  = str(row.get('Container Name', '')).strip()
        if cas and cas not in ('nan', '', '-'):
            if cas not in lib:
                lib[cas] = {'name': nm, 'smiles': smi if smi not in ('nan','Unknown','') else '', 'formula': frm}
            elif not lib[cas]['smiles'] and smi not in ('nan','Unknown',''):
                lib[cas]['smiles'] = smi
    return lib

# ── Build unified reaction table ──────────────────────────────────────────────
def build_reactions():
    csv = pd.read_csv(IN_CSV)
    rows = []
    for _, r in csv.iterrows():
        atype = str(r['amine_type']).strip().lower()
        sub   = str(r['substituent']).strip()
        pos   = str(r['position']).strip().lower()

        # Normalise substituent / position
        sub_map = {'none':'none','f':'F','cl':'Cl','br':'Br','oh':'OH','ome':'OMe',
                   'cf3':'CF3','me':'Me','ch3':'Me','tolyl':'Me'}
        pos_map = {'none':'none','para':'para','meta':'meta','ortho':'ortho',
                   'p':'para','m':'meta','o':'ortho'}
        sub_norm = sub_map.get(sub.lower(), sub)
        pos_norm = pos_map.get(pos.lower(), pos)

        # Determine paper source from compound_id
        cid_str = str(r.get('compound_id', '')).strip()
        if cid_str.startswith('P3'):
            paper_id = 'paper_3'; paper_src = 'SynOpen 2023 — Long et al. (Paper 3, biological)'
        elif cid_str.startswith('P2'):
            paper_id = 'paper_2'; paper_src = 'SynOpen 2023 — Long et al. (Paper 2, heterocyclic)'
        elif cid_str == 'FAILED':
            paper_id = 'paper_1'; paper_src = 'SynOpen 2023 — Long et al. (hard fail)'
        else:
            paper_id = 'paper_1'; paper_src = 'SynOpen 2023 — Long et al.'

        key = (atype, sub_norm, pos_norm)
        if key in AMINE_MAP:
            am = AMINE_MAP[key]
        else:
            key2 = (atype, sub.lower(), pos_norm)
            am = AMINE_MAP.get(key2) or AMINE_MAP.get(('heterocyclic', sub.lower(), 'none'))
            if not am:
                print(f'  [SKIP] No SMILES mapping: {atype} / {sub} / {pos}  (id={cid_str})')
                continue

        # Mark known hard-fail reactions
        hard_fail = (cid_str == 'FAILED') or \
                    (atype == 'aniline' and sub_norm in ('Cl','Br') and pos_norm == 'ortho') or \
                    (atype == 'heterocyclic' and sub.lower() == 'piperazine')

        # Compute product via RDKit SMARTS (skip hard fails)
        product_smi = '' if hard_fail else (compute_product(am['smiles']) or '')

        rows.append({
            'dataset_version':       'v1.1',
            'batch_id':              'batch_001',
            'paper_id':              paper_id,
            'paper_source':          paper_src,
            'reaction_id':           f"R{cid_str.zfill(3) if cid_str.isdigit() else cid_str}",
            'hard_fail':             hard_fail,
            'squarate_name':         SQUARATE_NAME,
            'squarate_smiles':       SQUARATE_SMILES,
            'squarate_cas':          SQUARATE_CAS,
            'amine_name':            am['name'],
            'amine_smiles':          am['smiles'],
            'amine_cas':             am['cas'],
            'solvent':               str(r.get('solvent','')).strip(),
            'temperature_celsius':   25 if str(r.get('temperature','')).strip().lower() in ('r.t.','rt','room temp') else str(r.get('temperature','')),
            'time_h':                r.get('time_h', ''),
            'catalyst':              str(r.get('catalyst','')).strip(),
            'catalyst_pct':          r.get('catalyst_loading_mol_pct', ''),
            'product_smiles':        product_smi,
            'yield_percent':         r.get('yield_pct', ''),
            '_amine_smiles_raw':     am['smiles'],
        })
    print(f'Built {len(rows)} reaction rows from CSV')
    return rows

# ── Enrich with PubChem ───────────────────────────────────────────────────────
def enrich(rows, library):
    # Load cache
    cache = {}
    if os.path.exists(CACHE_F):
        with open(CACHE_F, 'r') as f:
            cache = json.load(f)
        print(f'Loaded {len(cache)} cached PubChem entries')

    # Collect unique SMILES to look up
    all_smiles = set()
    for row in rows:
        all_smiles.add(row['squarate_smiles'])
        all_smiles.add(row['amine_smiles'])
        if row['product_smiles']:
            all_smiles.add(row['product_smiles'])
    uncached = [s for s in all_smiles if canonicalise(s) and canonicalise(s) not in cache]
    print(f'Fetching {len(uncached)} new SMILES from PubChem (cached: {len(all_smiles)-len(uncached)})...')

    for i, smi in enumerate(uncached):
        print(f'  [{i+1}/{len(uncached)}] {smi[:50]}...' if len(smi)>50 else f'  [{i+1}/{len(uncached)}] {smi}')
        pubchem_by_smiles(smi, cache)

    # Save cache
    with open(CACHE_F, 'w') as f:
        json.dump(cache, f, indent=2)
    print(f'Cache saved ({len(cache)} entries)')

    enriched = []
    for row in rows:
        sq_pc  = pubchem_by_smiles(row['squarate_smiles'], cache)
        am_pc  = pubchem_by_smiles(row['amine_smiles'],    cache)
        pr_pc  = pubchem_by_smiles(row['product_smiles'],  cache) if row['product_smiles'] else {}

        # RDKit-computed formula (local, always available)
        sq_formula = formula_from_smiles(row['squarate_smiles'])
        am_formula = formula_from_smiles(row['amine_smiles'])
        pr_formula = formula_from_smiles(row['product_smiles']) if row['product_smiles'] else ''

        # SMILES match checks (our SMILES vs PubChem canonical)
        sq_match = smiles_match(row['squarate_smiles'], sq_pc.get('pubchem_smiles','')) if sq_pc.get('pubchem_smiles') else 'UNVERIFIED'
        am_match = smiles_match(row['amine_smiles'],    am_pc.get('pubchem_smiles','')) if am_pc.get('pubchem_smiles') else 'UNVERIFIED'

        # Library cross-check (by CAS)
        sq_in_lib = row['squarate_cas'] in library
        am_in_lib = row['amine_cas']    in library

        # CAS validation: does our CAS match PubChem's CAS?
        sq_cas_ok = (row['squarate_cas'] == sq_pc.get('pubchem_cas','')) if sq_pc.get('pubchem_cas') else 'UNVERIFIED'
        am_cas_ok = (row['amine_cas']    == am_pc.get('pubchem_cas','')) if am_pc.get('pubchem_cas') else 'UNVERIFIED'

        # Overall QC score (0-10)
        score = 0
        if row['amine_smiles']:           score += 2
        if row['product_smiles']:         score += 2
        if am_match is True:             score += 2
        if sq_match is True:             score += 1
        if am_in_lib:                    score += 1
        if row['yield_percent'] != '':   score += 1
        if am_pc.get('cid'):             score += 1

        notes = []
        if am_match is False:   notes.append('SMILES mismatch vs PubChem')
        if sq_match is False:   notes.append('Squarate SMILES mismatch')
        if am_cas_ok is False:  notes.append('CAS mismatch vs PubChem')
        if not am_in_lib:       notes.append('Amine not in ChemInventory')
        if not row['product_smiles']: notes.append('Product SMILES missing')

        enriched.append({
            # ── Core identifiers ──
            'dataset_version':        row['dataset_version'],
            'batch_id':               row['batch_id'],
            'paper_id':               row['paper_id'],
            'paper_source':           row['paper_source'],
            'reaction_id':            row['reaction_id'],
            'hard_fail':              'YES' if row.get('hard_fail') else 'NO',

            # ── Squarate (SM1) ──
            'sq_name':                row['squarate_name'],
            'sq_smiles':              row['squarate_smiles'],
            'sq_cas':                 row['squarate_cas'],
            'sq_formula_rdkit':       sq_formula or '',
            'sq_formula_pubchem':     sq_pc.get('pubchem_formula', ''),
            'sq_iupac_pubchem':       sq_pc.get('pubchem_iupac', ''),
            'sq_smiles_match':        'MATCH' if sq_match is True else ('MISMATCH' if sq_match is False else 'UNVERIFIED'),
            'sq_cas_pubchem':         sq_pc.get('pubchem_cas', ''),
            'sq_cas_match':           'MATCH' if sq_cas_ok is True else ('MISMATCH' if sq_cas_ok is False else 'UNVERIFIED'),
            'sq_in_library':          'YES' if sq_in_lib else 'NO',
            'sq_pubchem_cid':         sq_pc.get('cid', ''),

            # ── Amine (SM2) ──
            'am_name':                row['amine_name'],
            'am_smiles':              row['amine_smiles'],
            'am_cas':                 row['amine_cas'],
            'am_formula_rdkit':       am_formula or '',
            'am_formula_pubchem':     am_pc.get('pubchem_formula', ''),
            'am_iupac_pubchem':       am_pc.get('pubchem_iupac', ''),
            'am_smiles_match':        'MATCH' if am_match is True else ('MISMATCH' if am_match is False else 'UNVERIFIED'),
            'am_cas_pubchem':         am_pc.get('pubchem_cas', ''),
            'am_cas_match':           'MATCH' if am_cas_ok is True else ('MISMATCH' if am_cas_ok is False else 'UNVERIFIED'),
            'am_in_library':          'YES' if am_in_lib else 'NO',
            'am_pubchem_cid':         am_pc.get('cid', ''),

            # ── Conditions ──
            'solvent':                row['solvent'],
            'temperature_celsius':    row['temperature_celsius'],
            'time_h':                 row['time_h'],
            'catalyst':               row['catalyst'],
            'catalyst_pct':           row['catalyst_pct'],

            # ── Product ──
            'product_smiles':         row['product_smiles'],
            'product_smiles_source':  'RDKit SMARTS' if row['product_smiles'] else 'MISSING',
            'product_formula_rdkit':  pr_formula or '',
            'product_iupac_pubchem':  pr_pc.get('pubchem_iupac', ''),
            'product_pubchem_cid':    pr_pc.get('cid', ''),

            # ── Yield ──
            'yield_percent':          row['yield_percent'],

            # ── Quality ──
            'qc_score_10':            score,
            'qc_pct':                 round(score * 10, 0),
            'qc_status':              'READY' if score >= 8 else 'REVIEW' if score >= 5 else 'INCOMPLETE',
            'qc_notes':               '; '.join(notes) if notes else 'OK',
        })
    return enriched

# ── Write Excel v1.1 ───────────────────────────────────────────────────────────
def write_excel(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Dataset_v1.1'

    # Styles
    hdr_font  = Font(name='Arial', bold=True, color='FFFFFF', size=9)
    hdr_fill  = PatternFill('solid', start_color='1F3864')
    grp_fill  = PatternFill('solid', start_color='2C4A7C')   # sub-header
    ok_fill   = PatternFill('solid', start_color='E8F5E9')
    rev_fill  = PatternFill('solid', start_color='FFF8E1')
    inc_fill  = PatternFill('solid', start_color='FFEBEE')
    alt_fill  = PatternFill('solid', start_color='F5F5F5')
    match_fill= PatternFill('solid', start_color='E8F5E9')
    mis_fill  = PatternFill('solid', start_color='FFEBEE')
    uv_fill   = PatternFill('solid', start_color='FFF9C4')
    center    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin      = Border(
        left=Side(style='thin', color='D0D0D0'), right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),  bottom=Side(style='thin', color='D0D0D0'),
    )

    # Column definitions: (key, header, width, align)
    COLS = [
        # Identifiers
        ('dataset_version',     'Version',            9,  'c'),
        ('batch_id',            'Batch',              10, 'c'),
        ('paper_id',            'Paper',              9,  'c'),
        ('paper_source',        'Source',             28, 'l'),
        ('reaction_id',         'Rxn ID',             9,  'c'),
        ('hard_fail',           'Hard Fail',          10, 'c'),
        # SM1
        ('sq_name',             'Squarate Name',      22, 'l'),
        ('sq_smiles',           'Squarate SMILES',    30, 'l'),
        ('sq_cas',              'Sq CAS',             12, 'c'),
        ('sq_formula_rdkit',    'Sq Formula',         12, 'c'),
        ('sq_iupac_pubchem',    'Sq IUPAC (PubChem)', 30, 'l'),
        ('sq_smiles_match',     'Sq SMILES Match',    14, 'c'),
        ('sq_cas_match',        'Sq CAS Match',       12, 'c'),
        ('sq_in_library',       'Sq In Library',      12, 'c'),
        ('sq_pubchem_cid',      'Sq PubChem CID',     13, 'c'),
        # SM2
        ('am_name',             'Amine Name',         22, 'l'),
        ('am_smiles',           'Amine SMILES',       30, 'l'),
        ('am_cas',              'Amine CAS',          12, 'c'),
        ('am_formula_rdkit',    'Amine Formula',      12, 'c'),
        ('am_iupac_pubchem',    'Amine IUPAC (PubChem)', 30, 'l'),
        ('am_smiles_match',     'Amine SMILES Match', 14, 'c'),
        ('am_cas_match',        'Amine CAS Match',    12, 'c'),
        ('am_in_library',       'Amine In Library',   12, 'c'),
        ('am_pubchem_cid',      'Amine PubChem CID',  14, 'c'),
        # Conditions
        ('solvent',             'Solvent',            10, 'c'),
        ('temperature_celsius', 'Temp (C)',            9, 'c'),
        ('time_h',              'Time (h)',            9, 'c'),
        ('catalyst',            'Catalyst',           16, 'l'),
        ('catalyst_pct',        'Cat. mol%',           9, 'c'),
        # Product
        ('product_smiles',      'Product SMILES',     36, 'l'),
        ('product_smiles_source','Product Source',    14, 'c'),
        ('product_formula_rdkit','Product Formula',   13, 'c'),
        ('product_iupac_pubchem','Product IUPAC',     30, 'l'),
        # Yield & QC
        ('yield_percent',       'Yield (%)',           9, 'c'),
        ('qc_score_10',         'QC Score /10',       12, 'c'),
        ('qc_pct',              'QC %',                8, 'c'),
        ('qc_status',           'Status',             12, 'c'),
        ('qc_notes',            'QC Notes',           40, 'l'),
    ]

    # Header row
    for ci, (key, hdr, width, align) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=hdr)
        c.font      = hdr_font
        c.fill      = hdr_fill
        c.alignment = center
        c.border    = thin
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 30

    match_cols  = {'sq_smiles_match', 'sq_cas_match', 'am_smiles_match', 'am_cas_match'}
    lib_cols    = {'sq_in_library', 'am_in_library'}

    for ri, row in enumerate(rows, 2):
        status = row.get('qc_status', 'REVIEW')
        base_fill = ok_fill if status == 'READY' else rev_fill if status == 'REVIEW' else inc_fill
        if ri % 2 == 1 and status == 'READY': base_fill = alt_fill

        for ci, (key, _, _, align) in enumerate(COLS, 1):
            val = row.get(key, '')
            c   = ws.cell(row=ri, column=ci, value=val)
            c.font      = Font(name='Arial', size=9)
            c.border    = thin
            c.alignment = center if align == 'c' else left

            # Special cell colouring
            if key in match_cols:
                c.fill = match_fill if val == 'MATCH' else mis_fill if val == 'MISMATCH' else uv_fill
                c.font = Font(name='Arial', size=9, bold=True,
                              color='1B5E20' if val=='MATCH' else 'B71C1C' if val=='MISMATCH' else '5D4037')
            elif key in lib_cols:
                c.fill = match_fill if val == 'YES' else uv_fill
            elif key == 'qc_status':
                c.fill = ok_fill if val=='READY' else rev_fill if val=='REVIEW' else inc_fill
                c.font = Font(name='Arial', size=9, bold=True)
            else:
                c.fill = base_fill

        ws.row_dimensions[ri].height = 20

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}1'

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ss = wb.create_sheet('Summary_v1.1')
    ss.column_dimensions['A'].width = 32
    ss.column_dimensions['B'].width = 55

    n_ready    = sum(1 for r in rows if r['qc_status'] == 'READY')
    n_review   = sum(1 for r in rows if r['qc_status'] == 'REVIEW')
    n_inc      = sum(1 for r in rows if r['qc_status'] == 'INCOMPLETE')
    n_match    = sum(1 for r in rows if r['am_smiles_match'] == 'MATCH')
    n_mismatch = sum(1 for r in rows if r['am_smiles_match'] == 'MISMATCH')
    n_lib      = sum(1 for r in rows if r['am_in_library']   == 'YES')
    n_product  = sum(1 for r in rows if r['product_smiles'])

    summary_rows = [
        ('CHEMDEM DATASET v1.1 — ENRICHMENT SUMMARY', ''),
        ('', ''),
        ('Total reactions',          len(rows)),
        ('SMILES verified vs PubChem (match)', n_match),
        ('SMILES mismatch vs PubChem',  n_mismatch),
        ('Amine in ChemInventory',   n_lib),
        ('Product SMILES computed',  n_product),
        ('', ''),
        ('QC RESULTS', ''),
        ('READY (score ≥ 8/10)',     n_ready),
        ('REVIEW (score 5–7/10)',    n_review),
        ('INCOMPLETE (score < 5)',   n_inc),
        ('', ''),
        ('ENRICHMENT SOURCES', ''),
        ('Local SMILES mapping',     'table1_reactions.csv + AMINE_MAP lookup'),
        ('Product SMILES',           'RDKit SMARTS reaction (auto-computed)'),
        ('IUPAC names',              'PubChem REST API (compound/smiles/…/IUPACName)'),
        ('Molecular formulas',       'RDKit CalcMolFormula + PubChem cross-check'),
        ('CAS validation',           'PubChem synonyms endpoint'),
        ('Library cross-check',      'ChemInventory-Export.xlsx (lecturer inventory)'),
        ('', ''),
        ('HOW TO READ MATCH COLUMNS', ''),
        ('MATCH',       'Our SMILES = PubChem canonical SMILES (after RDKit canonicalisation)'),
        ('MISMATCH',    'Different structure — check manually, may need correction'),
        ('UNVERIFIED',  'PubChem had no entry or lookup failed'),
    ]

    SECTION_KEYS = {'CHEMDEM DATASET v1.1 — ENRICHMENT SUMMARY','QC RESULTS','ENRICHMENT SOURCES','HOW TO READ MATCH COLUMNS'}
    for ri2, (label, val) in enumerate(summary_rows, 1):
        ca = ss.cell(row=ri2, column=1, value=label)
        cb = ss.cell(row=ri2, column=2, value=val)
        if label in SECTION_KEYS:
            for c in (ca, cb):
                c.font  = Font(name='Arial', bold=True, size=10, color='FFFFFF')
                c.fill  = hdr_fill
                c.alignment = left
        elif label:
            ca.font = Font(name='Arial', bold=True, size=10)
            cb.font = Font(name='Arial', size=10)
        ca.border = thin; cb.border = thin

    wb.save(OUT_XL)
    print(f'\nSaved: {OUT_XL}')

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print('=== Chemdem Dataset Enrichment v1.1 ===\n')

    print('Loading ChemInventory library...')
    library = load_library()
    print(f'Library: {len(library)} unique CAS entries\n')

    print('Building reaction table from CSV...')
    rows = build_reactions()
    print()

    print('Enriching with PubChem...')
    enriched = enrich(rows, library)
    print()

    print('Writing Excel...')
    write_excel(enriched)

    # Quick stats
    print('\n=== SUMMARY ===')
    for status in ('READY', 'REVIEW', 'INCOMPLETE'):
        n = sum(1 for r in enriched if r['qc_status'] == status)
        print(f'  {status}: {n}/{len(enriched)}')
    matches   = sum(1 for r in enriched if r['am_smiles_match'] == 'MATCH')
    mismatches= sum(1 for r in enriched if r['am_smiles_match'] == 'MISMATCH')
    print(f'  PubChem SMILES match: {matches}, mismatch: {mismatches}')
    print('\nDone.')
